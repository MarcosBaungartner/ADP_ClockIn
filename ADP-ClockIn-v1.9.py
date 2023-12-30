#!/usr/bin/env python
# ADP-ClockIn
# Developed by Marcos Roberto Baungartner
# v1.9 Final

import glob
import os
import time
import tkinter
import warnings
from datetime import date, datetime, timedelta

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

warnings.simplefilter("ignore")

# Cria path.txt se nao existir
if not os.path.isfile(f'{os.getcwd()}\\path.txt'):
    with open('path.txt', 'w', encoding='utf_8') as arq: # Cria arquivo path.txt e abre em modo escrita
        arq.write('# Download path\n')
        arq.write(f'download_path={os.getcwd()}') # This path
        arq.write('# Username\n')
        arq.write('username=\n')
        arq.write('# Password\n')
        arq.write('password=')
        arq.close() # Fecha arquivo

with open(f'{os.getcwd()}\\path.txt', encoding='utf_8') as arq: # Abre arquivo em modo somente leitura
    lines = arq.readlines() # Le linha por linha
    arq.close() # Fecha arquivo

# Percorre linhas de path.txt
for line in lines:
    if line.find('download_path=') >= 0: # Se encontrar variavel
        download_path = f'{(line.rsplit("download_path=", 1)[-1]).strip()}\\{date.today().year}' # Concatena valor da variavel + ano corrente
        if not os.path.isdir(download_path): # Se diretorio do ano corrente nao existir
            os.mkdir(download_path) # Cria diretorio
    if line.find('username=') >= 0: # Se encontrar variavel
        username_txt = (line.rsplit('username=', 1)[-1]).strip() # Recebe valor da variavel
    if line.find('password=') >= 0: # Se encontrar variavel
        password_txt = (line.rsplit('password=', 1)[-1]).strip() # Recebe valor da variavel


root = tkinter.Tk()


def main():
    """Funcao Principal"""
    login()


def login():
    """Janela de Login para ADP"""
    root.title("Acesso ADP")
    root.geometry('300x140')
    root.eval('tk::PlaceWindow . center')

    tkinter.Label(root, text="Usuario:").place(x=20, y=15)
    entry_user = tkinter.Entry(root, width=30)  # Text box do login
    entry_user.insert(0, username_txt)
    entry_user.place(x=80, y=15)

    tkinter.Label(root, text='Senha:').place(x=20, y=40)
    entry_password = tkinter.Entry(root, show="*", width=30) # Text box da senha
    entry_password.insert(0, password_txt)
    entry_password.place(x=80, y=40)

    chk_box = tkinter.IntVar()
    entry_check = tkinter.Checkbutton(root,text='Baixar Demonstrativo', variable=chk_box, onvalue=1, offvalue=0) # Checkbutton
    entry_check.select()
    entry_check.place(x=80, y=60)

    button = tkinter.Button(root, width=10, text='OK', command=lambda: login_adp(entry_user.get(), entry_password.get(), chk_box.get()))
    button.place(x=110, y=95)

    root.bind('<Return>', (lambda event: login_adp(entry_user.get(), entry_password.get(), chk_box.get())))

    entry_password.focus()

    root.mainloop()



def marcar_ponto(driver, mes_ano, dia_mes, entrada1, saida1, entrada2, saida2, forcar_marcacao, desfazer):
    """Marca o ponto no ADP e retorna Status"""
    try:

        # Encontra data da tabela atual no ADP
        mes_ano_adp_array = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="js-react-app"]/div/div/div[4]/div[2]/div[2]/div/div/div[1]/div/div/div/p/time[1]'))).get_attribute('innerHTML').split('/')
        mes_ano_array = mes_ano.split('/') # Armazena mes e ano no array

        xpath_dia_mes = '//*[@id="js-react-app"]/div/div/div[4]/div[2]/div[2]/div/div/div[2]/div/div[2]/div[2]/table/tbody/' # XPATH do elemento Dia/Mes
        xpath_desc = '//*[@id="js-react-app"]/div/div/div[4]/div[2]/div[2]/div/div/div[2]/div/div[2]/div[2]/table/tbody/' # XPATH do elemento Descricao

        # Se Mes/Ano da planilha for menor do que no ADP, retrocede o Mes
        if int(mes_ano_array[0]) < int(mes_ano_adp_array[1]) or int(mes_ano_array[1]) < int(mes_ano_adp_array[2]):
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="js-react-app"]/div/div/div[4]/div[2]/div[2]/div/div/div[1]/div/div/a'))).click()

        # Se Mes/Ano da planilha for maior do que no ADP, avanca o Mes
        elif int(mes_ano_array[0]) > int(mes_ano_adp_array[1]) or int(mes_ano_array[1]) > int(mes_ano_adp_array[2]):
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="js-react-app"]/div/div/div[4]/div[2]/div[2]/div/div/div[1]/div/div/a[2]'))).click()

        # Se Mes/Ano da planilha for menor que mes corrente, variaveis recebem mes anterior
        if int(mes_ano_array[0]) < date.today().month or int(mes_ano_array[1]) < date.today().year:
            xpath_dia_mes = '//*[@id="js-react-app"]/div/div/div[4]/div[2]/div[2]/div/div/div[2]/div/div/div[2]/table/tbody/' # XPATH do elemento Dia/Mes
            xpath_desc = '//*[@id="js-react-app"]/div/div/div[4]/div[2]/div[2]/div/div/div[2]/div/div/div[2]/table/tbody/' # XPATH do elemento Descricao

        # Verifica tipo de dia: normal, folga, DSR, recesso ou feriado
        for dia in range(1,32):
            if WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, f'{xpath_dia_mes}tr[%i]/td[1]/div/strong' %dia))).get_attribute('innerHTML') == dia_mes:
                marcacao = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, f'{xpath_desc}tr[%i]/td[2]/div/div' %dia))).get_attribute('innerHTML')
                if len(marcacao) > 20:
                    marcacao = ''
                break

        # Se o dia for especial e nao houver 'MARCAR' na celula da coluna F
        if marcacao != '' and not forcar_marcacao:
            return marcacao # Retorna o tipo do dia

        if desfazer:
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, f'//*[normalize-space()="{dia_mes}"]'))).click()
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//button[normalize-space()="Desfazer"]'))).click()
            return 'DESFEITO'


        ###########################################################################################################################################################
        ################### REALIZA MARCACAO DE PONTO #############################################################################################################
        ###########################################################################################################################################################

        # Clica no ponto escolhido (dia/mes)
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, f'//*[normalize-space()="{dia_mes}"]'))).click()

        # Clica no Botao Sugerir ajuste
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//button[normalize-space()="Sugerir ajuste"]'))).click()

        # Clica no Botao Inserir Marcações
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//button[normalize-space()="Inserir Marcações"]'))).click()



        # Primeira Hora
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//time[@data-testid="txt_timesheet_adjustments_modal-marking-item-0-time"]'))).click()

        # Justificativa
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//input[@name="justify"]'))).send_keys('Horário de trabalho')

        # hh:mm
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="js-react-portal"]/div/div[1]/section/div/div/div[4]/div/section/div[2]/ol/li[1]/div[2]/div/div/div/div/div/div[1]/div[2]/div/label/div/input'))).send_keys(entrada1)



        # Segunda Hora
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//time[@data-testid="txt_timesheet_adjustments_modal-marking-item-1-time"]'))).click()

        # hh:mm
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="js-react-portal"]/div/div[1]/section/div/div/div[4]/div/section/div[2]/ol/li[2]/div[2]/div/div/div/div/div/div[1]/div[2]/div/label/div/input'))).send_keys(saida1)


        if entrada2 is not None:
            # Terceira Hora
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//time[@data-testid="txt_timesheet_adjustments_modal-marking-item-2-time"]'))).click()

            # hh:mm
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="js-react-portal"]/div/div[1]/section/div/div/div[4]/div/section/div[2]/ol/li[3]/div[2]/div/div/div/div/div/div[1]/div[2]/div/label/div/input'))).send_keys(entrada2)



            # Quarta Hora
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//time[@data-testid="txt_timesheet_adjustments_modal-marking-item-3-time"]'))).click()

            # hh:mm
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="js-react-portal"]/div/div[1]/section/div/div/div[4]/div/section/div[2]/ol/li[4]/div[2]/div/div/div/div/div/div[1]/div[2]/div/label/div/input'))).send_keys(saida2)



        # Clica no Botao Salvar ajuste
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="js-react-portal"]/div/div[1]/section/div/div/div[4]/div/section/div[2]/button'))).click()

        # Clica no Botao Voltar Para Lista
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="js-react-portal"]/div/div[1]/section/div/footer/div[2]/div/div/div/div/button'))).click()

    except Exception:
        return 'ERRO'
    else:
        return 'OK'



def login_adp(username, password, chk_box):
    """Acessa ADP, marca o ponto, e preenche planilha"""

    root.destroy()

    if username == '' or password == '':
        return # Se login estiver vazio, termina funcao

    try: #Configuracao do chromedriver
        chrome_options = Options()
        chrome_options.add_argument("start-maximized")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])
        chrome_options.add_experimental_option("prefs",{"download.default_directory" : download_path})
        driver = webdriver.Chrome(options=chrome_options)
    except Exception as err:
        print (f'Erro: Verifique se a versao do chromedriver confere com a do Google Chrome. Argumentos {err.args}')
        input('Pressione Enter para finalizar...')
    else:
        try: #Sessão do ADP
            driver.get('https://expert.brasil.adp.com/ipclogin/1/loginform.html?TYPE=33554433&REALMOID=06-000a1470-e058-1656-b22f-441e0bf0d04d&GUID=&SMAUTHREASON=0&METHOD=GET&SMAGENTNAME=jKCbBo2iXmPsA0rq7iagICXFbbwYt9UvfBpgtMIzDmy9OoXI6rviphUJYrlLzFqY&TARGET=-SM-https%3a%2f%2fexpert%2ebrasil%2eadp%2ecom%2fexpert%2f')

            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="login"]'))).send_keys(username) #Login ADP
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="login-pw"]'))).send_keys(password) #Senha ADP
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div/div/section[1]/div/form/div[3]/div/button'))).click() # Botao Entrar

            if chk_box == 1: # Se Checkbutton 'Baixar Demonstrativo' estiver selecionado
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="js-react-app"]/div/div/div[4]/div[1]/nav[1]/a[3]'))).click() #Menu Pagamentos
                try:
                    for i in range(1,3):
                        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, f'//*[@id="js-react-app"]/div/div/div[4]/div[2]/div[2]/div[1]/div/div/div[1]/ol/li[{i}]/a/div'))).click() # Acessa pagamento
                        nome_pagamento = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, f'//*[@id="js-react-app"]/div/div/div[4]/div[2]/div[2]/div[1]/div/div/div[1]/ol/li[{i}]/a/div/h3/strong'))).get_attribute('innerHTML') # Recebe nome do pagamento
                        data_pagamento = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, f'//*[@id="js-react-app"]/div/div/div[4]/div[2]/div[2]/div[1]/div/div/div[1]/ol/li[{i}]/a/div/dl/dd[2]'))).get_attribute('innerHTML').split('/') # Recebe data de pagamento
                        nome_pagamento = f'{data_pagamento[1]}-{data_pagamento[0]}_{nome_pagamento}.pdf' # Concatena data de pagamento com nome do pagamento

                        flag_arquivo_existe = False
                        arquivos = glob.glob(f'{download_path}\\*.pdf')
                        # Se houver mais de um arquivo xlsx no diretorio
                        if len(arquivos) > 1:
                            # Para cada arquivo no diretorio
                            for arquivo in arquivos:
                                # Se nome arquivo atual for diferente de relatorio-visita.xlsx
                                if arquivo == f'{download_path}\\{nome_pagamento}':
                                    flag_arquivo_existe = True
                                    break
                        if not flag_arquivo_existe:
                            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="js-react-app"]/div/div/div[4]/div[2]/div[2]/div[1]/div/div/div[2]/div/div[1]/h2/a/button[2]'))).click() # Download Demonstrativo
                            time.sleep(3) #Aguarda 3 segundos
                            os.rename(f'{download_path}\\Demonstrativo de Pagamento.pdf', f'{download_path}\\{nome_pagamento}') # Renomeia arquivo baixado
                except Exception as err:
                    pass

            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="js-react-app"]/div/div/div[4]/div[1]/nav[1]/a[2]'))).click() # Menu Ponto

        except Exception as err: # Se houve erro
            driver.quit() # Encerra driver
            print(f'Erro ao abrir sessao no portal ADP. Argumentos {err.args}')
            input('Pressione Enter para finalizar...')
        else:
            try: # Se nao houve erro
                wb = load_workbook(filename='ADP.xlsx', read_only=False) # Carrega Workbook
                ws = wb.active # Seta Worksheet

                mes = date.today().month + 16 # Posicao da coluna do mes da tabela da linha BANCO
                ws.cell(row=4, column=mes, value=WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="js-react-app"]/div/div/div[4]/div[2]/div[2]/div[1]/div/div[2]/aside/div[1]/div/div[3]/div/strong/span[1]'))).get_attribute('innerHTML')) # Celula recebe horas acumuladas do mes

                ws['N2'].value = date.today() # N2 recebe data da ultima atualizacao

                print('\n +============================================+')
                print(' |       DATA      | HORAS | EXTRA |  STATUS  |')

                flag_count = 0 # Inicializa variavel
                for row in range(1,ws.max_row + 1): # Percorre linhas da planilha
                    if ws['H' + str(row)].value is None or ws['H' + str(row)].value == 'MARCAR' or ws['H' + str(row)].value == 'DESFAZER': # Se Status estiver vazio ou MARCAR
                        data = ws['A' + str(row)].value
                        if data.date() < date.today(): # Se data da planilha for menor que data de hoje
                            try:
                                if data.date().day == 1:
                                    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="js-react-app"]/div/div/div[4]/div[2]/div[2]/div/div/div[1]/div/div/a'))).click() # Retorna ao mes anterior
                                    mes = 12 if date.today().month - 1 == 0 else date.today().month - 1 # Mes anterior
                                    mes = mes + 16 # Posicao da coluna do mes da tabela da linha ACUMULADO
                                else:
                                    mes = date.today().month + 16 # Posicao da coluna do mes da tabela da linha ACUMULADO
                                    flag_count = flag_count + 1 # Flag para atualizar o ACUMULADO/BANCO/FOLHA do mes apenas uma vez

                                if flag_count < 2:

                                    hour_minute = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="js-react-app"]/div/div/div[4]/div[2]/div[2]/div[1]/div/div[2]/aside/div[1]/div/div[2]/ul/li[1]/div/div/button/div/div[2]/span[1]'))).get_attribute('innerHTML')
                                    if (WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="js-react-app"]/div/div/div[4]/div[2]/div[2]/div[1]/div/div[2]/aside/div[1]/div/div[2]/ul/li[1]/div/div/button/div/div[2]/span[1]'))).get_attribute('class')).find('blue') < 0:
                                            hour_minute = '-' + hour_minute
                                    else:
                                        hour_minute = '+' + hour_minute
                                    ws.cell(row=5, column=mes, value=hour_minute) # Celula recebe BANCO de horas do mes

                                    # hour_minute = (WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="js-react-app"]/div/div/div[4]/div[2]/div[2]/div[1]/div/div[2]/aside/div[1]/div/div[2]/ul/li[1]/div/div/button/div/div[2]/span[1]'))).get_attribute('innerHTML')).split(':')
                                    # ws.cell(row=5, column=mes, value=timedelta(hours=int(hour_minute[0]), minutes=int(hour_minute[1]))) # Celula recebe BANCO de horas do mes
                                    # ws.cell(row=5, column=mes).number_format='[hh]:mm;@'
                                    ws.cell(row=5, column=mes).alignment = Alignment(horizontal='center')


                                    hour_minute = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="js-react-app"]/div/div/div[4]/div[2]/div[2]/div[1]/div/div[2]/aside/div[1]/div/div[2]/ul/li[2]/div/div/button/div/div[2]/span[1]'))).get_attribute('innerHTML')
                                    if (WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="js-react-app"]/div/div/div[4]/div[2]/div[2]/div[1]/div/div[2]/aside/div[1]/div/div[2]/ul/li[2]/div/div/button/div/div[2]/span[1]'))).get_attribute('class')).find('blue') < 0:
                                        hour_minute = '-' + hour_minute
                                    else:
                                        hour_minute = '+' + hour_minute
                                    ws.cell(row=6, column=mes, value=hour_minute)

                                    # hour_minute = (WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="js-react-app"]/div/div/div[4]/div[2]/div[2]/div[1]/div/div[2]/aside/div[1]/div/div[2]/ul/li[2]/div/div/button/div/div[2]/span[1]'))).get_attribute('innerHTML')).split(':')
                                    # ws.cell(row=6, column=mes, value=timedelta(hours=int(hour_minute[0]), minutes=int(hour_minute[1]))) # Celula recebe FOLHA de horas do mes
                                    # ws.cell(row=6, column=mes).number_format='[hh]:mm;@'
                                    ws.cell(row=6, column=mes).alignment = Alignment(horizontal='center')

                                    ws.cell(row=4, column=mes, value=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="js-react-app"]/div/div/div[4]/div[2]/div[2]/div[1]/div/div[2]/aside/div[1]/div/div[3]/div/strong/span[1]'))).get_attribute('innerHTML')) # Celula recebe horas acumuladas do mes

                                    wb.save('ADP.xlsx') # Salva planilha

                                if mes < date.today().month + 16:
                                    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="js-react-app"]/div/div/div[4]/div[2]/div[2]/div/div/div[1]/div/div/a[2]'))).click() # Avança mes
                            except Exception as err:
                                WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="js-react-app"]/div/div/div[4]/div[2]/div[2]/div/div/div[1]/div/div/a[2]'))).click() # Avança mes
                            finally:
                                print(' |-----------------+-------+-------+----------|')
                                mes_ano = (ws['A' + str(row)].value).strftime('%m/%Y') # mm/yyyy
                                dia_mes = data.strftime('%d/%m') # dd/mm
                                entrada1 = (ws['B' + str(row)].value).strftime('%H:%M')
                                saida1 = (ws['C' + str(row)].value).strftime('%H:%M')
                                e1 = ws['B' + str(row)].value
                                s1 = ws['C' + str(row)].value
                                total = timedelta(hours=s1.hour, minutes=s1.minute) - timedelta(hours=e1.hour, minutes=e1.minute)
                                if ws['D' + str(row)].value is None:
                                    entrada2 = None
                                    saida2 = None
                                else:
                                    entrada2 = (ws['D' + str(row)].value).strftime('%H:%M')
                                    saida2 = (ws['E' + str(row)].value).strftime('%H:%M')
                                    e2 = ws['D' + str(row)].value
                                    s2 = ws['E' + str(row)].value
                                    total = timedelta(seconds=total.seconds) + timedelta(hours=s2.hour, minutes=s2.minute) - timedelta(hours=e2.hour, minutes=e2.minute)
                                extra = timedelta(seconds=0) if timedelta(hours=8) > total else total-timedelta(hours=8)
                                forcar_marcacao = True if ws['H' + str(row)].value == 'MARCAR' else False
                                desfazer = True if ws['H' + str(row)].value == 'DESFAZER' else False
                                result = marcar_ponto(driver, mes_ano, dia_mes, entrada1, saida1, entrada2, saida2, forcar_marcacao, desfazer) # Marca o ponto
                                if result == 'OK':
                                    print(f' | {data.strftime("%a")}, {data.date()} | {datetime.strptime(str(total), "%H:%M:%S").strftime("%H:%M")} | {datetime.strptime(str(extra), "%H:%M:%S").strftime("%H:%M")} | {result.center(8)} |')
                                    ws['F' + str(row)].value = total # Se ponto marcado, marca HORA TOTAL na celula TOTAL
                                    ws['G' + str(row)].value = extra # Se ponto marcado, marca HORA EXTRA na celula EXTRA
                                else:
                                    print(f' | {data.strftime("%a")}, {data.date()} | {"     "} | {"     "} | {result.center(8)} |')
                                if result != 'ERRO': # Se não houver ERRO
                                    ws['H' + str(row)].value = result # Se ponto marcado, marca OK/FOLGA/DSR/FERIADO/ABONO/DESFEITO na celula Status
                                    wb.save('ADP.xlsx') # Salva planilha
                            if result == 'ERRO': # Se houver ERRO
                                break # termina rotina
                        else: # Se data da planilha nao for menor que data de hoje
                            break # termina rotina

                print(' +============================================+')
                wb.close() # Fecha planilha
                driver.quit() # Encerra driver
                input('\n\nPressione Enter para finalizar...')

            except Exception as err:
                driver.quit() # Encerra driver
                print (f'Erro: Verifique se a planilha ADP.xlsm existe no diretorio raiz. Argumentos {err.args}')
                input('Pressione Enter para finalizar...')

# PRINCIPAL
main()
